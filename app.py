import io
import os
import re
import tempfile
from datetime import datetime

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sqlalchemy import or_

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "nlp-recruitment-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", "mysql+pymysql://root:@localhost:3306/nlp_recruitment"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5 MB upload limit

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {"pdf", "docx", "txt"}

db = SQLAlchemy(app)

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)  # plain text by request
    role = db.Column(db.String(20), default="user")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resumes = db.relationship("Resume", backref="user", lazy=True)
    applications = db.relationship("Application", backref="user", lazy=True)


class Job(db.Model):
    __tablename__ = "jobs"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    location = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    applications = db.relationship("Application", backref="job", lazy=True)


class Resume(db.Model):
    __tablename__ = "resumes"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=True)
    content = db.Column(db.Text, nullable=False)
    skills = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    applications = db.relationship("Application", backref="resume", lazy=True)


class Application(db.Model):
    __tablename__ = "applications"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    job_id = db.Column(db.Integer, db.ForeignKey("jobs.id"), nullable=False)
    resume_id = db.Column(db.Integer, db.ForeignKey("resumes.id"), nullable=True)
    status = db.Column(db.String(40), default="Under Review")
    match_score = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ---------------------------------------------------------------------------
# NLP helpers
# ---------------------------------------------------------------------------

COMMON_SKILLS = [
    "python", "java", "javascript", "typescript", "c++", "c#", "ruby", "go",
    "rust", "swift", "kotlin", "php", "sql", "nosql", "mongodb", "postgresql",
    "mysql", "redis", "elasticsearch", "html", "css", "react", "angular", "vue",
    "node", "express", "django", "flask", "fastapi", "spring", "docker",
    "kubernetes", "aws", "azure", "gcp", "git", "ci/cd", "linux", "rest",
    "graphql", "microservices", "agile", "scrum", "machine learning", "deep learning",
    "nlp", "natural language processing", "tensorflow", "pytorch", "pandas",
    "numpy", "scikit-learn", "data analysis", "data science", "statistics",
    "project management", "leadership", "communication", "teamwork",
    "problem solving", "critical thinking", "excel", "power bi", "tableau",
    "figma", "photoshop", "ui/ux", "devops", "terraform", "ansible",
]


def extract_skills(text: str) -> list[str]:
    text_lower = text.lower()
    found = []
    for skill in COMMON_SKILLS:
        pattern = r"\b" + re.escape(skill).replace(r"\ ", r"\s+") + r"\b"
        if re.search(pattern, text_lower):
            found.append(skill)
    return sorted(set(found))


def score_resume_against_job(resume_text: str, job_description: str) -> float:
    """
    Blended NLP score: 40% TF-IDF cosine similarity + 60% skill overlap.
    Skill overlap is weighted more because short job descriptions dilute
    pure TF-IDF when compared against long resumes.
    """
    if not resume_text.strip() or not job_description.strip():
        return 0.0

    tfidf_score = 0.0
    try:
        vectorizer = TfidfVectorizer(stop_words="english", max_features=5000)
        tfidf_matrix = vectorizer.fit_transform([resume_text, job_description])
        tfidf_score = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
    except ValueError:
        pass

    resume_skills = set(extract_skills(resume_text))
    job_skills = set(extract_skills(job_description))
    if job_skills:
        skill_score = len(resume_skills & job_skills) / len(job_skills)
    else:
        skill_score = tfidf_score

    blended = (tfidf_score * 0.4) + (skill_score * 0.6)
    return round(blended * 100, 2)


def parse_resume_file(file_storage) -> str:
    """Extract plain text from uploaded PDF, DOCX, or TXT file."""
    filename = file_storage.filename.lower()
    if filename.endswith(".txt"):
        return file_storage.read().decode("utf-8", errors="replace")

    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name

    try:
        if filename.endswith(".pdf"):
            return _parse_pdf(tmp_path)
        elif filename.endswith(".docx"):
            return _parse_docx(tmp_path)
        return ""
    finally:
        os.unlink(tmp_path)


def _parse_pdf(path: str) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n".join(pages)


def _parse_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

def validate_name(name: str) -> str | None:
    if not name:
        return "Name is required."
    if not re.match(r"^[A-Za-z\s]+$", name):
        return "Name must contain only letters and spaces (no numbers or special characters)."
    if len(name) < 2:
        return "Name must be at least 2 characters."
    return None


def validate_username(username: str) -> str | None:
    if not username:
        return "Username is required."
    if " " in username:
        return "Username must not contain spaces."
    if username[0].isdigit():
        return "Username must not start with a number."
    if not re.match(r"^[A-Za-z][A-Za-z0-9_]*$", username):
        return "Username can only contain letters, numbers, and underscores."
    if len(username) < 3:
        return "Username must be at least 3 characters."
    return None


def validate_email(email: str) -> str | None:
    if not email:
        return "Email is required."
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    if not re.match(pattern, email):
        return "Please enter a valid email address."
    return None


def validate_password(password: str) -> str | None:
    if not password:
        return "Password is required."
    if len(password) < 6:
        return "Password must be at least 6 characters."
    return None

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def ensure_admin_user():
    admin = User.query.filter_by(email="admin@system.local").first()
    if not admin:
        admin = User(
            name="Admin",
            username="admin",
            email="admin@system.local",
            password="admin",
            role="admin",
        )
        db.session.add(admin)
        db.session.commit()


def current_user():
    if "user_id" not in session:
        return None
    return db.session.get(User, session["user_id"])


def login_required(role=None):
    def decorator(fn):
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                flash("Please log in to continue.", "warning")
                return redirect(url_for("login"))
            if role and user.role != role:
                flash("You are not authorized to access this area.", "danger")
                return redirect(url_for("home"))
            return fn(*args, **kwargs)
        wrapper.__name__ = fn.__name__
        return wrapper
    return decorator


def setup():
    db.create_all()
    ensure_admin_user()


with app.app_context():
    setup()

# ---------------------------------------------------------------------------
# Routes – public
# ---------------------------------------------------------------------------

@app.route("/")
def home():
    jobs = Job.query.order_by(Job.created_at.desc()).limit(3).all()
    return render_template("home.html", jobs=jobs, user=current_user())


@app.route("/about")
def about():
    return render_template("about.html", user=current_user())


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        errors = []
        name_err = validate_name(name)
        if name_err:
            errors.append(name_err)
        username_err = validate_username(username)
        if username_err:
            errors.append(username_err)
        email_err = validate_email(email)
        if email_err:
            errors.append(email_err)
        password_err = validate_password(password)
        if password_err:
            errors.append(password_err)

        if errors:
            for e in errors:
                flash(e, "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(email=email).first():
            flash("An account with that email already exists.", "danger")
            return redirect(url_for("register"))
        if User.query.filter_by(username=username).first():
            flash("That username is already taken.", "danger")
            return redirect(url_for("register"))

        user = User(name=name, username=username, email=email, password=password, role="user")
        db.session.add(user)
        db.session.commit()

        session["user_id"] = user.id
        session["role"] = user.role
        flash("Welcome! You are registered and logged in.", "success")
        return redirect(url_for("user_dashboard"))

    return render_template("register.html", user=current_user())


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip().lower()
        password = request.form.get("password", "")

        if identifier == "admin" and password == "admin":
            admin_user = User.query.filter_by(role="admin").first()
            if admin_user:
                session["user_id"] = admin_user.id
                session["role"] = "admin"
                flash("Logged in as admin.", "success")
                return redirect(url_for("admin_dashboard"))

        user = User.query.filter(
            or_(
                User.email == identifier,
                User.username == identifier,
                User.name.ilike(identifier),
            )
        ).first()
        if not user or user.password != password:
            flash("Invalid credentials. Try again.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = user.id
        session["role"] = user.role
        flash("Welcome back!", "success")
        if user.role == "admin":
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("user_dashboard"))

    return render_template("login.html", user=current_user())


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("home"))

# ---------------------------------------------------------------------------
# Routes – candidate dashboard
# ---------------------------------------------------------------------------

@app.route("/dashboard")
@login_required(role="user")
def user_dashboard():
    user = current_user()
    jobs = Job.query.order_by(Job.created_at.desc()).all()
    applications = (
        Application.query.filter_by(user_id=user.id)
        .order_by(Application.created_at.desc())
        .all()
    )
    latest_resume = (
        Resume.query.filter_by(user_id=user.id)
        .order_by(Resume.created_at.desc())
        .first()
    )
    return render_template(
        "dashboard_user.html",
        user=user,
        jobs=jobs,
        applications=applications,
        latest_resume=latest_resume,
    )


@app.route("/upload_resume", methods=["POST"])
@login_required(role="user")
def upload_resume():
    user = current_user()
    content = ""
    filename = None

    if "resume_file" in request.files:
        file = request.files["resume_file"]
        if file and file.filename and allowed_file(file.filename):
            filename = file.filename
            content = parse_resume_file(file)
        elif file and file.filename and not allowed_file(file.filename):
            flash("Unsupported file type. Please upload PDF, DOCX, or TXT.", "danger")
            return redirect(url_for("user_dashboard"))

    if not content:
        content = request.form.get("resume_text", "").strip()

    if not content:
        flash("Please upload a file or paste your resume text.", "warning")
        return redirect(url_for("user_dashboard"))

    skills = extract_skills(content)
    resume = Resume(
        user_id=user.id,
        content=content,
        filename=filename,
        skills=", ".join(skills) if skills else None,
    )
    db.session.add(resume)
    db.session.commit()

    skill_msg = f" Detected skills: {', '.join(skills)}." if skills else ""
    flash(f"Resume saved.{skill_msg}", "success")
    return redirect(url_for("user_dashboard"))


@app.route("/apply/<int:job_id>", methods=["POST"])
@login_required(role="user")
def apply(job_id):
    user = current_user()
    job = Job.query.get_or_404(job_id)

    existing = Application.query.filter_by(user_id=user.id, job_id=job.id).first()
    if existing:
        flash("You have already applied to this job.", "warning")
        return redirect(url_for("user_dashboard"))

    content = ""
    filename = None

    if "resume_file" in request.files:
        file = request.files["resume_file"]
        if file and file.filename and allowed_file(file.filename):
            filename = file.filename
            content = parse_resume_file(file)

    if not content:
        content = request.form.get("resume_text", "").strip()

    latest_resume = (
        Resume.query.filter_by(user_id=user.id)
        .order_by(Resume.created_at.desc())
        .first()
    )

    if not content and not latest_resume:
        flash("Upload a resume before applying.", "warning")
        return redirect(url_for("user_dashboard"))

    if content:
        skills = extract_skills(content)
        resume = Resume(
            user_id=user.id,
            content=content,
            filename=filename,
            skills=", ".join(skills) if skills else None,
        )
        db.session.add(resume)
        db.session.flush()
    else:
        resume = latest_resume

    score = score_resume_against_job(resume.content, job.description)
    application = Application(
        user_id=user.id,
        job_id=job.id,
        resume_id=resume.id,
        status="Under Review",
        match_score=score,
    )
    db.session.add(application)
    db.session.commit()
    flash(f"Applied to {job.title}. NLP match score: {score}%", "success")
    return redirect(url_for("user_dashboard"))

# ---------------------------------------------------------------------------
# Routes – admin
# ---------------------------------------------------------------------------

@app.route("/admin")
@login_required(role="admin")
def admin_dashboard():
    jobs = Job.query.order_by(Job.created_at.desc()).all()
    applications = Application.query.order_by(Application.match_score.desc()).all()
    total_users = User.query.filter(User.role == "user").count()
    total_jobs = Job.query.count()
    total_apps = Application.query.count()
    return render_template(
        "dashboard_admin.html",
        user=current_user(),
        jobs=jobs,
        applications=applications,
        total_users=total_users,
        total_jobs=total_jobs,
        total_apps=total_apps,
    )


@app.route("/admin/jobs", methods=["POST"])
@login_required(role="admin")
def create_job():
    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    location = request.form.get("location", "").strip()
    if not title or not description:
        flash("Title and description are required.", "warning")
        return redirect(url_for("admin_dashboard"))

    job = Job(title=title, description=description, location=location)
    db.session.add(job)
    db.session.commit()
    flash("Job posted.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/jobs/<int:job_id>/edit", methods=["GET", "POST"])
@login_required(role="admin")
def edit_job(job_id):
    job = Job.query.get_or_404(job_id)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        location = request.form.get("location", "").strip()
        if not title or not description:
            flash("Title and description are required.", "warning")
            return redirect(url_for("edit_job", job_id=job.id))
        job.title = title
        job.description = description
        job.location = location
        db.session.commit()
        flash(f'Job "{job.title}" updated.', "success")
        return redirect(url_for("admin_dashboard"))
    return render_template("edit_job.html", user=current_user(), job=job)


@app.route("/admin/jobs/<int:job_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_job(job_id):
    job = Job.query.get_or_404(job_id)
    Application.query.filter_by(job_id=job.id).delete()
    db.session.delete(job)
    db.session.commit()
    flash(f'Job "{job.title}" deleted.', "info")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/applications/<int:application_id>/status", methods=["POST"])
@login_required(role="admin")
def update_application_status(application_id):
    status = request.form.get("status", "Under Review")
    application = Application.query.get_or_404(application_id)
    application.status = status
    db.session.commit()
    flash(f"Application {application.id} marked as {status}.", "info")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/reports")
@login_required(role="admin")
def reports():
    job_stats = (
        db.session.query(Job.title, db.func.count(Application.id))
        .outerjoin(Application)
        .group_by(Job.id)
        .all()
    )
    status_stats = (
        db.session.query(Application.status, db.func.count(Application.id))
        .group_by(Application.status)
        .all()
    )
    avg_score = db.session.query(db.func.avg(Application.match_score)).scalar() or 0
    top_candidates = (
        db.session.query(User.name, db.func.avg(Application.match_score).label("avg_score"))
        .join(Application)
        .filter(User.role == "user")
        .group_by(User.id)
        .order_by(db.text("avg_score DESC"))
        .limit(5)
        .all()
    )
    apps_over_time = (
        db.session.query(
            db.func.date(Application.created_at).label("date"),
            db.func.count(Application.id),
        )
        .group_by(db.func.date(Application.created_at))
        .order_by(db.text("date"))
        .all()
    )
    skill_frequency = _compute_skill_frequency()
    return render_template(
        "reports.html",
        user=current_user(),
        job_stats=job_stats,
        status_stats=status_stats,
        avg_score=round(avg_score, 2),
        top_candidates=top_candidates,
        apps_over_time=apps_over_time,
        skill_frequency=skill_frequency,
    )


def _compute_skill_frequency() -> list[tuple[str, int]]:
    resumes = Resume.query.filter(Resume.skills.isnot(None)).all()
    freq: dict[str, int] = {}
    for r in resumes:
        for skill in r.skills.split(", "):
            s = skill.strip().lower()
            if s:
                freq[s] = freq.get(s, 0) + 1
    return sorted(freq.items(), key=lambda x: x[1], reverse=True)[:15]


def _gather_report_data():
    """Collect all report data into plain lists for export."""
    applications = (
        db.session.query(
            User.name,
            Job.title,
            Application.match_score,
            Application.status,
            Application.created_at,
        )
        .join(User, Application.user_id == User.id)
        .join(Job, Application.job_id == Job.id)
        .order_by(Application.match_score.desc())
        .all()
    )
    job_stats = (
        db.session.query(Job.title, db.func.count(Application.id))
        .outerjoin(Application)
        .group_by(Job.id)
        .all()
    )
    status_stats = (
        db.session.query(Application.status, db.func.count(Application.id))
        .group_by(Application.status)
        .all()
    )
    top_candidates = (
        db.session.query(User.name, db.func.avg(Application.match_score).label("avg"))
        .join(Application)
        .filter(User.role == "user")
        .group_by(User.id)
        .order_by(db.text("avg DESC"))
        .limit(10)
        .all()
    )
    skill_frequency = _compute_skill_frequency()
    return applications, job_stats, status_stats, top_candidates, skill_frequency


@app.route("/admin/reports/download/excel")
@login_required(role="admin")
def download_report_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    applications, job_stats, status_stats, top_candidates, skill_frequency = _gather_report_data()
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 1 – All applications
    ws = wb.active
    ws.title = "Applications"
    style_header(ws, ["Candidate", "Job", "Match Score (%)", "Status", "Applied On"])
    for r, app in enumerate(applications, 2):
        ws.cell(row=r, column=1, value=app[0]).border = thin_border
        ws.cell(row=r, column=2, value=app[1]).border = thin_border
        ws.cell(row=r, column=3, value=float(app[2])).border = thin_border
        ws.cell(row=r, column=4, value=app[3]).border = thin_border
        ws.cell(row=r, column=5, value=app[4].strftime("%Y-%m-%d %H:%M")).border = thin_border
    auto_width(ws)

    # Sheet 2 – Applications per job
    ws2 = wb.create_sheet("Per Job")
    style_header(ws2, ["Job Title", "Applications"])
    for r, (title, count) in enumerate(job_stats, 2):
        ws2.cell(row=r, column=1, value=title).border = thin_border
        ws2.cell(row=r, column=2, value=count).border = thin_border
    auto_width(ws2)

    # Sheet 3 – Status distribution
    ws3 = wb.create_sheet("Status")
    style_header(ws3, ["Status", "Count"])
    for r, (status, count) in enumerate(status_stats, 2):
        ws3.cell(row=r, column=1, value=status).border = thin_border
        ws3.cell(row=r, column=2, value=count).border = thin_border
    auto_width(ws3)

    # Sheet 4 – Top candidates
    ws4 = wb.create_sheet("Top Candidates")
    style_header(ws4, ["Candidate", "Avg Match Score (%)"])
    for r, (name, avg) in enumerate(top_candidates, 2):
        ws4.cell(row=r, column=1, value=name).border = thin_border
        ws4.cell(row=r, column=2, value=round(float(avg), 1)).border = thin_border
    auto_width(ws4)

    # Sheet 5 – Skill frequency
    ws5 = wb.create_sheet("Skills")
    style_header(ws5, ["Skill", "Count"])
    for r, (skill, count) in enumerate(skill_frequency, 2):
        ws5.cell(row=r, column=1, value=skill).border = thin_border
        ws5.cell(row=r, column=2, value=count).border = thin_border
    auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"recruitment_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )


@app.route("/admin/reports/download/pdf")
@login_required(role="admin")
def download_report_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    applications, job_stats, status_stats, top_candidates, skill_frequency = _gather_report_data()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Recruitment Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 8 * mm))

    header_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    # Applications table
    elements.append(Paragraph("All Applications (ranked by score)", styles["Heading2"]))
    data = [["Candidate", "Job", "Score %", "Status", "Date"]]
    for app in applications:
        data.append([app[0], app[1], f"{app[2]:.1f}", app[3], app[4].strftime("%Y-%m-%d")])
    if len(data) > 1:
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle(header_style))
        elements.append(t)
    else:
        elements.append(Paragraph("No applications yet.", styles["Normal"]))
    elements.append(Spacer(1, 6 * mm))

    # Applications per job
    elements.append(Paragraph("Applications per Job", styles["Heading2"]))
    data = [["Job Title", "Applications"]]
    for title, count in job_stats:
        data.append([title, str(count)])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(header_style))
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Status distribution
    elements.append(Paragraph("Status Distribution", styles["Heading2"]))
    data = [["Status", "Count"]]
    for status, count in status_stats:
        data.append([status, str(count)])
    if len(data) > 1:
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle(header_style))
        elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Top candidates
    elements.append(Paragraph("Top Candidates", styles["Heading2"]))
    data = [["Candidate", "Avg Score %"]]
    for name, avg in top_candidates:
        data.append([name, f"{float(avg):.1f}"])
    if len(data) > 1:
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle(header_style))
        elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Skill frequency
    if skill_frequency:
        elements.append(Paragraph("Most Common Skills", styles["Heading2"]))
        data = [["Skill", "Count"]]
        for skill, count in skill_frequency:
            data.append([skill, str(count)])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle(header_style))
        elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"recruitment_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True)
